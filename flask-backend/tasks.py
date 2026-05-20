import os
import time
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from services.reconciliation import (
    process_receivements,
    process_payments,
    process_summary
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def process_files_task(ac_path, ob_path, output_path, timestamp, use_ai=False, ai_provider="gemini", log_queue=None):
    start_time = time.time()
    ai_ran = False
    logger.info(f"Job started for output: {output_path}")
    try:
        recv_df = process_receivements(ac_path, ob_path, timestamp)
        pay_df = process_payments(ac_path, ob_path, timestamp)
        summary_df = process_summary(ac_path, ob_path)

        with pd.ExcelWriter(output_path, engine='openpyxl') as w:
            recv_df.to_excel(w, sheet_name='Receivements', index=False)
            pay_df.to_excel(w, sheet_name='Payments', index=False)
            summary_df.to_excel(
                w,
                sheet_name='Summary',
                index=True,
                index_label=['Date', 'Metric']
            )

        wb = load_workbook(output_path)
        grey = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        for ws in wb.worksheets:
            max_col = ws.max_column
            for i in range(2, ws.max_row + 1):
                if ws[f"A{i}"].value is None:
                    for col in ws.iter_cols(min_col=1, max_col=max_col, min_row=i, max_row=i):
                        col[0].fill = grey

        wb.save(output_path)

        if use_ai:
            if ai_provider == "gemini":
                api_key  = os.getenv("GEMINI_API_KEY")
                key_name = "GEMINI_API_KEY"
            else:
                api_key  = (os.getenv("OPEN_ROUTER_API_KEY") or "").strip()
                key_name = "OPEN_ROUTER_API_KEY"

            if not api_key:
                logger.warning(f"AI matching requested but {key_name} is not set")
            else:
                def log(msg):
                    if log_queue:
                        log_queue.put(msg)
                try:
                    from services.matcher import apply_highlights
                    apply_highlights(output_path, api_key=api_key, provider=ai_provider, log=log)
                    ai_ran = True
                except Exception as e:
                    logger.warning(f"AI highlighting failed: {e}")

        duration = time.time() - start_time
        logger.info(f"Job completed in {duration:.2f} seconds")

    except Exception as e:
        logger.exception(f"Job failed with exception: {e}")

    finally:
        if os.path.exists(ac_path):
            os.remove(ac_path)
        if os.path.exists(ob_path):
            os.remove(ob_path)

    return {"path": output_path, "ai_ran": ai_ran}
