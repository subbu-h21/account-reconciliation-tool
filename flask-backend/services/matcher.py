import json
import logging
import asyncio
import re
from datetime import datetime, date as date_type

import networkx as nx
from openai import AsyncOpenAI
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

logger = logging.getLogger(__name__)

YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
BLUE   = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
GREY   = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
WHITE  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

MODEL          = "openai/gpt-4o-mini"
PILE_TOLERANCE = 1.0

# 0-based column indices: Date=0, Books name=1, Books amt=2, blank=3, Bank name=4, Bank amt=5
BOOKS_NAME_COL = 1
BOOKS_AMT_COL  = 2
BANK_NAME_COL  = 4
BANK_AMT_COL   = 5

_TOOL = {
    "type": "function",
    "function": {
        "name": "submit_pile_matches",
        "description": (
            "Submit one-to-one matching decisions for this pile. "
            "Every row ID must appear exactly once across matches, "
            "unmatched_books, and unmatched_bank."
        ),
        "parameters": {
            "type": "object",
            "required": ["reasoning", "matches", "unmatched_books", "unmatched_bank"],
            "properties": {
                "reasoning": {
                    "type": "string",
                    "description": (
                        "Think step by step BEFORE committing to decisions. "
                        "Reason about each candidate pair before writing your final answer."
                    ),
                },
                "matches": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "required": ["books_id", "bank_id", "reason", "confidence"],
                        "properties": {
                            "books_id":   {"type": "string"},
                            "bank_id":    {"type": "string"},
                            "reason":     {"type": "string"},
                            "confidence": {"type": "number", "minimum": 0.0, "maximum": 1.0},
                        },
                    },
                },
                "unmatched_books": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "required": ["books_id", "reason"],
                        "properties": {
                            "books_id": {"type": "string"},
                            "reason":   {"type": "string"},
                        },
                    },
                },
                "unmatched_bank": {
                    "type": "array",
                    "items": {
                        "type": "object",
                        "required": ["bank_id", "reason"],
                        "properties": {
                            "bank_id": {"type": "string"},
                            "reason":  {"type": "string"},
                        },
                    },
                },
            },
        },
    },
}

_SYSTEM_PROMPT = (
    "You are a bank reconciliation assistant matching entries between a bank statement "
    "and internal accounting books for the same date.\n\n"
    "All rows in a pile have amounts within ₹1 of each other — do NOT re-evaluate money. "
    "Focus entirely on whether the names refer to the same real-world entity or transaction.\n\n"
    "Common equivalences to recognise:\n"
    "  - 'NEFT-KAPILA-PHARMA' = 'Kapila Pharma'\n"
    "  - 'salary to manju' = 'manjunath'\n"
    "  - Prefix/suffix noise on bank side (NEFT-, MClick/, UBI-, etc.) — strip and compare core name\n"
    "  - Abbreviations, spacing, case differences in the same entity name\n\n"
    "Rules:\n"
    "  1. One-to-one: each books entry pairs with at most one bank entry and vice versa.\n"
    "  2. Every ID must appear exactly once across matches, unmatched_books, unmatched_bank.\n"
    "  3. Do NOT force pairings — leave entries unmatched if no credible name match exists.\n"
    "  4. A false match is worse than leaving a row unmatched."
)


def _date_str(val):
    if isinstance(val, datetime):
        return val.date().isoformat()
    if isinstance(val, date_type):
        return val.isoformat()
    return str(val)


def _read_sheet_by_date(ws):
    """
    Returns {date_value: {"books": [(name, amount, row_num, entry_id), ...],
                          "bank":  [(name, amount, row_num, entry_id), ...]}}
    entry_id is a unique int across all entries in the sheet.
    """
    groups = {}
    order  = []
    entry_counter = 0

    for row in ws.iter_rows(min_row=2):
        date_val = row[0].value
        if date_val is None:
            continue
        if date_val not in groups:
            groups[date_val] = {"books": [], "bank": []}
            order.append(date_val)

        row_num    = row[0].row
        books_name = row[BOOKS_NAME_COL].value
        books_amt  = row[BOOKS_AMT_COL].value
        bank_name  = row[BANK_NAME_COL].value
        bank_amt   = row[BANK_AMT_COL].value

        if books_name is not None:
            groups[date_val]["books"].append(
                (str(books_name), float(books_amt) if books_amt is not None else None, row_num, entry_counter)
            )
            entry_counter += 1
        if bank_name is not None:
            groups[date_val]["bank"].append(
                (str(bank_name), float(bank_amt) if bank_amt is not None else None, row_num, entry_counter)
            )
            entry_counter += 1

    # return in original date order
    return {d: groups[d] for d in order}


def _tier1_exact(books_entries, bank_entries):
    """
    Exact match: case-insensitive stripped name + amount rounded to 2dp.
    Returns (matched_pairs, remaining_books, remaining_bank).
    matched_pairs: list of (books_entry, bank_entry) tuples.
    """
    used_books = set()
    used_bank  = set()
    matched    = []

    for be in books_entries:
        bname, bamt, _, bid = be
        if bamt is None:
            continue
        bkey = (bname.strip().lower(), round(bamt, 2))
        for ke in bank_entries:
            kname, kamt, _, kid = ke
            if kid in used_bank or kamt is None:
                continue
            kkey = (kname.strip().lower(), round(kamt, 2))
            if bkey == kkey:
                matched.append((be, ke))
                used_books.add(bid)
                used_bank.add(kid)
                break

    rem_books = [e for e in books_entries if e[3] not in used_books]
    rem_bank  = [e for e in bank_entries  if e[3] not in used_bank]
    return matched, rem_books, rem_bank


def _build_piles(books_entries, bank_entries):
    """
    Connected components on ±₹1 amount tolerance.
    Returns list of {"books": [...], "bank": [...]} dicts.
    """
    G = nx.Graph()

    for e in books_entries:
        G.add_node(("B", e[3]), entry=e)
    for e in bank_entries:
        G.add_node(("K", e[3]), entry=e)

    for be in books_entries:
        bamt = be[1]
        if bamt is None:
            continue
        for ke in bank_entries:
            kamt = ke[1]
            if kamt is None:
                continue
            if abs(bamt - kamt) <= PILE_TOLERANCE:
                G.add_edge(("B", be[3]), ("K", ke[3]))

    piles = []
    for comp in nx.connected_components(G):
        b_entries = [G.nodes[n]["entry"] for n in comp if n[0] == "B"]
        k_entries = [G.nodes[n]["entry"] for n in comp if n[0] == "K"]
        piles.append({"books": b_entries, "bank": k_entries})

    return piles


def _pile_prompt(pile):
    lines = [
        "Bank reconciliation — AI matching pass.",
        "",
        "CONTEXT:",
        "  All rows are from the same date.",
        "  Amounts are within ₹1 of each other (that is why they are in this pile).",
        "  Decide IDENTITY only — do the names refer to the same real-world entity?",
        "  One-to-one constraint: each entry may be used in at most one match.",
        "  Every ID (B1..Bn and K1..Km) must appear exactly once across",
        "  matches / unmatched_books / unmatched_bank.",
        "",
        "BOOKS ROWS (internal accounting):",
    ]
    for i, (name, amt, _, _id) in enumerate(pile["books"], 1):
        lines.append(f"  B{i}: name={name}  amount={amt}")

    lines += ["", "BANK ROWS (bank statement):"]
    for i, (name, amt, _, _id) in enumerate(pile["bank"], 1):
        lines.append(f"  K{i}: name={name}  amount={amt}")

    lines += [
        "",
        "Call submit_pile_matches.",
        "Every Bi and every Kj must appear exactly once. Use B1..Bn and K1..Km as the IDs.",
    ]
    return "\n".join(lines)


async def _ai_pile(client, pile):
    if not pile["books"] or not pile["bank"]:
        return None

    b_id_map = {f"B{i}": e for i, e in enumerate(pile["books"], 1)}
    k_id_map = {f"K{i}": e for i, e in enumerate(pile["bank"],  1)}

    prompt_text = _pile_prompt(pile)

    try:
        response = await client.chat.completions.create(
            model=MODEL,
            temperature=0,
            tools=[_TOOL],
            tool_choice={"type": "function", "function": {"name": "submit_pile_matches"}},
            messages=[
                {"role": "system", "content": _SYSTEM_PROMPT},
                {"role": "user",   "content": prompt_text},
            ],
        )
    except Exception as exc:
        logger.warning("AI pile call failed: %s", exc)
        return None

    raw = json.loads(response.choices[0].message.tool_calls[0].function.arguments)
    logger.debug("AI pile response: %s", raw.get("reasoning", "")[:120])

    seen_b: set[str] = set()
    seen_k: set[str] = set()
    matched_pairs = []

    for m in raw.get("matches", []):
        bid = m.get("books_id", "")
        kid = m.get("bank_id",  "")

        if bid not in b_id_map or kid not in k_id_map:
            logger.warning("AI returned unknown IDs %s/%s — skipping", bid, kid)
            continue
        if bid in seen_b or kid in seen_k:
            logger.warning("AI returned duplicate IDs %s/%s — skipping", bid, kid)
            continue

        be = b_id_map[bid]
        ke = k_id_map[kid]
        # amount tolerance check
        if be[1] is not None and ke[1] is not None and abs(be[1] - ke[1]) > PILE_TOLERANCE:
            logger.warning("AI match %s/%s fails amount tolerance — skipping", bid, kid)
            continue

        seen_b.add(bid)
        seen_k.add(kid)
        matched_pairs.append((be, ke))

    return matched_pairs


def _fuzzy_pair(books_entries, bank_entries):
    """
    Jaccard word-set similarity on names. Returns list of
    (books_entry_or_None, bank_entry_or_None) pairs.
    """
    def _words(s):
        return set(re.sub(r"[^a-z0-9 ]", " ", s.strip().lower()).split())

    candidates = []
    for bi, be in enumerate(books_entries):
        bw = _words(be[0])
        if not bw:
            continue
        for ki, ke in enumerate(bank_entries):
            kw = _words(ke[0])
            if not kw:
                continue
            union = bw | kw
            if not union:
                continue
            score = len(bw & kw) / len(union)
            if score > 0:
                candidates.append((score, bi, ki))

    candidates.sort(reverse=True)

    used_b: set[int] = set()
    used_k: set[int] = set()
    pairs = []

    for score, bi, ki in candidates:
        if bi not in used_b and ki not in used_k:
            pairs.append((books_entries[bi], bank_entries[ki]))
            used_b.add(bi)
            used_k.add(ki)

    for bi, be in enumerate(books_entries):
        if bi not in used_b:
            pairs.append((be, None))
    for ki, ke in enumerate(bank_entries):
        if ki not in used_k:
            pairs.append((None, ke))

    return pairs


def _rewrite_sheet(ws, date_blocks):
    """
    Delete all data rows, then rewrite in order:
    yellow (T1) → blue (T2 AI) → white (T3 fuzzy/unmatched) → grey separator per date.
    """
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    row_idx = 2
    for block in date_blocks:
        date_val = block["date"]

        def write_row(books_entry, bank_entry, fill):
            nonlocal row_idx
            bname = books_entry[0] if books_entry else None
            bamt  = books_entry[1] if books_entry else None
            kname = bank_entry[0]  if bank_entry  else None
            kamt  = bank_entry[1]  if bank_entry  else None

            vals = [date_val, bname, bamt, None, kname, kamt]
            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill = fill
            row_idx += 1

        for be, ke in block["t1_pairs"]:
            write_row(be, ke, YELLOW)

        for be, ke in block["ai_pairs"]:
            write_row(be, ke, BLUE)

        for be, ke in block["t3_pairs"]:
            write_row(be, ke, WHITE)

        # grey separator
        for col_idx in range(1, 7):
            cell = ws.cell(row=row_idx, column=col_idx, value=None)
            cell.fill = GREY
        row_idx += 1


def apply_matching(output_path, api_key, log=None):
    def emit(msg):
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {msg}"
        logger.debug(line)
        if log:
            log(line)

    emit("AI matching started")

    wb = load_workbook(output_path)

    for sheet_name in ["Receivements", "Payments"]:
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]
        emit(f"Processing sheet: {sheet_name}")

        date_data = _read_sheet_by_date(ws)
        date_blocks = []

        for date_val, sides in date_data.items():
            t1_pairs, rem_books, rem_bank = _tier1_exact(sides["books"], sides["bank"])
            piles = _build_piles(rem_books, rem_bank)

            ai_piles   = [p for p in piles if p["books"] and p["bank"]]
            solo_piles = [p for p in piles if not p["books"] or not p["bank"]]

            solo_books = [e for p in solo_piles if not p["bank"]  for e in p["books"]]
            solo_bank  = [e for p in solo_piles if not p["books"] for e in p["bank"]]

            date_blocks.append({
                "date":       date_val,
                "t1_pairs":   t1_pairs,
                "ai_piles":   ai_piles,
                "solo_books": solo_books,
                "solo_bank":  solo_bank,
                "ai_pairs":   [],
                "t3_pairs":   [],
            })
            emit(f"  {_date_str(date_val)}: {len(t1_pairs)} exact, {len(ai_piles)} piles to AI, {len(solo_books)+len(solo_bank)} solo")

        # fire all AI piles concurrently
        indexed_piles = [(i, p) for i, block in enumerate(date_blocks) for p in block["ai_piles"]]

        if indexed_piles:
            emit(f"Sending {len(indexed_piles)} piles to AI ({MODEL})...")
            client = AsyncOpenAI(api_key=api_key, base_url="https://openrouter.ai/api/v1")

            async def run_all():
                tasks = [_ai_pile(client, pile) for _, pile in indexed_piles]
                return await asyncio.gather(*tasks, return_exceptions=True)

            results = asyncio.run(run_all())

            for (block_idx, _), result in zip(indexed_piles, results):
                if isinstance(result, Exception):
                    emit(f"  pile error: {result}")
                elif result:
                    date_blocks[block_idx]["ai_pairs"].extend(result)

        # Tier 3 fuzzy for each date
        for block in date_blocks:
            ai_matched_book_ids = {e[3] for be, ke in block["ai_pairs"] for e in [be]}
            ai_matched_bank_ids = {e[3] for be, ke in block["ai_pairs"] for e in [ke]}

            unmatched_books = block["solo_books"] + [
                e for p in block["ai_piles"] for e in p["books"]
                if e[3] not in ai_matched_book_ids
            ]
            unmatched_bank = block["solo_bank"] + [
                e for p in block["ai_piles"] for e in p["bank"]
                if e[3] not in ai_matched_bank_ids
            ]
            block["t3_pairs"] = _fuzzy_pair(unmatched_books, unmatched_bank)

        _rewrite_sheet(ws, date_blocks)
        emit(f"Sheet {sheet_name} rewritten.")

    wb.save(output_path)
    emit("Done.")
